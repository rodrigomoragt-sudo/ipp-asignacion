import pandas as pd
import json
from datetime import datetime, timedelta
from collections import defaultdict
import calendar

class GeneradorPlanVisitas:
    def __init__(self):
        """
        Inicializar con datos de los excels desde carpeta datos/.
        Si algún archivo falta (ej. contenedor recién desplegado, sin datos
        cargados todavía) se propaga el FileNotFoundError tal cual — quien
        instancia esta clase decide qué hacer (app.py la atrapa y deja el
        servidor arriba con datos_cargados=False, para que el usuario pueda
        subir/sincronizar las tablas desde la UI).
        """
        self.df_cedis = pd.read_excel("datos/Clientes CEDIS.xlsx")
        self.df_ipp = pd.read_excel("datos/Clientes IPP Últimos 3 Meses.xlsx")  # Últimos 3 meses
        self.df_ipp_actual = pd.read_excel("datos/Clientes IPP Mes Actual.xlsx")  # Mes actual
        self.df_ef = pd.read_excel("datos/Clientes Equipo Frío.xlsx")

        # Crear conjuntos de clientes para búsqueda rápida
        self.clientes_ef_ids = set(self.df_ef['Cliente_id'].unique())
        self.clientes_ipp_dict = self._crear_dict_ipp()

        # Clientes IPP encuestados en el mes actual (del archivo actual)
        self.clientes_ipp_mes_actual = self._cargar_clientes_ipp_mes_actual()

    def _crear_dict_ipp(self):
        """Crear diccionario de clientes IPP con sus datos y fechas"""
        ipp_dict = {}
        for idx, row in self.df_ipp.iterrows():
            cod_cliente = row['Cod Cliente']
            ipp_dict[cod_cliente] = {
                'nombre': row.get('Nomb_cliente', ''),
                'zona': row.get('Zona', ''),
                'segmento': row.get('Segmento', '')
            }
        return ipp_dict

    def _cargar_clientes_ipp_mes_actual(self):
        """
        Cargar clientes que YA han sido encuestados en el mes actual
        Estos no deben ser incluidos en nuevas planificaciones de IPP
        Se cargan del archivo 'Clientes IPP_ actual.xlsx'
        """
        clientes_encuestados = set(self.df_ipp_actual['Cod Cliente'].unique())
        return clientes_encuestados

    def clasificar_cliente(self, codigo_cliente, segmento):
        """
        Clasificar cliente según PRIORIDADES CORRECTAS:
        1. Equipo Frío (EF)
        2. BLINDAR
        3. DESARROLLAR
        4. MANTENER
        5. OPTIMIZAR
        + EXCLUIR: Ya encuestados en mes actual
        """
        try:
            cod_int = int(codigo_cliente)
        except:
            cod_int = codigo_cliente

        # EXCLUSION: Si ya fue encuestado en el mes actual, EXCLUIR
        if cod_int in self.clientes_ipp_mes_actual:
            return 'Excluido', 99

        # Prioridad 1: EQUIPO FRÍO
        if cod_int in self.clientes_ef_ids:
            return 'Equipo Frío', 1

        # Clasificar por segmento
        if segmento and pd.notna(segmento):
            seg_lower = str(segmento).lower().strip()

            # Prioridad 2: BLINDAR
            if 'blind' in seg_lower or 'btl' in seg_lower:
                return 'Blindar', 2

            # Prioridad 3: DESARROLLAR
            elif 'desarro' in seg_lower:
                return 'Desarrollar', 3

            # Prioridad 4: MANTENER
            elif 'manten' in seg_lower:
                return 'Mantener', 4

            # Prioridad 5: OPTIMIZAR
            elif 'optim' in seg_lower:
                return 'Optimizar', 5

        # Por defecto: MANTENER
        return 'Mantener', 4

    def generar_plan(self, mes=7, ano=2026, inicio_visitas=5, fin_visitas=27, visitas_por_dia=20,
                     dias_exclusion=None, dias_sin_visita=None, supervisor=None):
        """
        Generar plan de visitas para un mes

        Args:
            mes: Número de mes (1-12)
            ano: Año
            inicio_visitas: Día del mes para comenzar visitas (1-31)
            fin_visitas: Día del mes para terminar visitas (1-31)
            visitas_por_dia: Número de visitas por supervisor por día
            dias_exclusion: Lista de días del mes sin visitas (ej: [28,29,30,31])
            dias_sin_visita: Lista de días de la semana sin visitas (0=Lun, 6=Dom)
            supervisor: Zona específica o None para todas
        """

        if dias_exclusion is None:
            dias_exclusion = []
        if dias_sin_visita is None:
            dias_sin_visita = [5, 6]  # Sábado y domingo por defecto

        # Obtener días hábiles del mes
        dias_habiles = self._obtener_dias_habiles(ano, mes, inicio_visitas, fin_visitas,
                                                   dias_exclusion, dias_sin_visita)

        # Agrupar clientes por supervisor
        clientes_por_supervisor = self._agrupar_clientes_por_supervisor()

        plan = {
            'mes': mes,
            'ano': ano,
            'inicio_visitas': inicio_visitas,
            'visitas_por_dia': visitas_por_dia,
            'dias_habiles': len(dias_habiles),
            'fecha_generacion': datetime.now().isoformat(),
            'supervisores': {}
        }

        # Generar plan para cada supervisor
        for zona in sorted(clientes_por_supervisor.keys()):
            if supervisor and str(supervisor) != str(zona):
                continue

            clientes = clientes_por_supervisor[zona]
            plan_supervisor = self._planificar_supervisor(
                zona, clientes, dias_habiles, visitas_por_dia
            )
            plan['supervisores'][str(zona)] = plan_supervisor

        return plan

    def _obtener_dias_habiles(self, ano, mes, inicio, fin, dias_exclusion, dias_sin_visita):
        """Obtener lista de fechas hábiles para el mes"""
        dias_habiles = []
        num_dias_mes = calendar.monthrange(ano, mes)[1]

        # Usar fin definido o hasta el último día del mes
        fecha_fin = min(fin, num_dias_mes)

        for dia in range(inicio, fecha_fin + 1):
            if dia in dias_exclusion:
                continue

            fecha = datetime(ano, mes, dia)
            if fecha.weekday() in dias_sin_visita:
                continue

            dias_habiles.append({
                'fecha': fecha.isoformat(),
                'dia': dia,
                'dia_semana': fecha.weekday(),
                'visitantes': 0,
                'clientes': []
            })

        return dias_habiles

    def _agrupar_clientes_por_supervisor(self):
        """Agrupar todos los clientes por supervisor"""
        supervisores = defaultdict(list)

        for idx, row in self.df_cedis.iterrows():
            codigo = row['Cliente: Código de Cliente']
            zona = row['Código Zona']
            segmento = row['Cliente: Segmento']
            ruta = row['Ruta']

            # Clasificar cliente
            clasificacion, prioridad = self.clasificar_cliente(int(codigo), segmento)

            cliente = {
                'codigo': int(codigo),
                'zona': int(zona),
                'ruta': str(ruta) if pd.notna(ruta) else 'SIN_RUTA',
                'segmento': str(segmento) if pd.notna(segmento) else '',
                'clasificacion': clasificacion,
                'prioridad': prioridad,
                'visitado_en_plan': False,
                'fecha_visita': None
            }

            supervisores[int(zona)].append(cliente)

        return supervisores

    def _planificar_supervisor(self, zona, clientes, dias_habiles, visitas_por_dia):
        """
        Planificar visitas para un supervisor
        CRITERIO: Cada DÍA una RUTA DIFERENTE rotando entre rutas
        """

        # Filtrar clientes excluidos
        clientes_validos = [c for c in clientes if c['prioridad'] < 99]
        clientes_excluidos = [c for c in clientes if c['prioridad'] >= 99]

        # Agrupar clientes por RUTA
        clientes_por_ruta = defaultdict(list)
        for cliente in clientes_validos:
            ruta = cliente['ruta'] if pd.notna(cliente['ruta']) else 'SIN_RUTA'
            clientes_por_ruta[ruta].append(cliente)

        # Ordenar clientes dentro de cada ruta por prioridad
        for ruta in clientes_por_ruta:
            clientes_por_ruta[ruta].sort(key=lambda x: x['prioridad'])

        # Crear lista de rutas rotables (CICLO)
        rutas_ciclo = list(clientes_por_ruta.keys())
        rutas_ciclo.sort()  # Ordenar alfabéticamente para consistencia

        plan_supervisor = {
            'zona': zona,
            'total_clientes': len(clientes),
            'clientes_validos': len(clientes_validos),
            'clientes_excluidos': len(clientes_excluidos),
            'clientes_planificados': 0,
            'total_rutas': len(rutas_ciclo),
            'cronograma': []
        }

        # Distribución por segmento
        resumen_segmentos = defaultdict(int)

        # Índices para cada ruta (para saber qué cliente tomar siguiente)
        cliente_idx_por_ruta = {ruta: 0 for ruta in rutas_ciclo}
        ruta_actual_idx = 0

        for dia_idx, dia_data in enumerate(dias_habiles):
            visitas_dia = []
            visitantes_asignados = 0

            # Obtener ruta actual en el ciclo
            if len(rutas_ciclo) > 0:
                ruta_actual = rutas_ciclo[ruta_actual_idx % len(rutas_ciclo)]
            else:
                continue

            # Asignar 20 clientes de la ruta actual
            clientes_en_ruta = clientes_por_ruta[ruta_actual]
            cliente_idx = cliente_idx_por_ruta[ruta_actual]

            while visitantes_asignados < visitas_por_dia and cliente_idx < len(clientes_en_ruta):
                cliente = clientes_en_ruta[cliente_idx]

                if not cliente['visitado_en_plan']:
                    visitas_dia.append({
                        'codigo_cliente': cliente['codigo'],
                        'ruta': cliente['ruta'],
                        'segmento': cliente['segmento'],
                        'clasificacion': cliente['clasificacion'],
                        'prioridad': cliente['prioridad']
                    })
                    cliente['visitado_en_plan'] = True
                    cliente['fecha_visita'] = dia_data['fecha']
                    resumen_segmentos[cliente['clasificacion']] += 1
                    visitantes_asignados += 1

                cliente_idx += 1

            # Guardar posición para próxima vez que visitemos esta ruta
            cliente_idx_por_ruta[ruta_actual] = cliente_idx

            if visitas_dia:
                plan_supervisor['cronograma'].append({
                    'fecha': dia_data['fecha'],
                    'dia': dia_data['dia'],
                    'dia_semana': ['Lun', 'Mar', 'Mie', 'Jue', 'Vie', 'Sab', 'Dom'][dia_data['dia_semana']],
                    'ruta': ruta_actual,
                    'visitas': visitas_dia,
                    'total_visitas': len(visitas_dia)
                })
                plan_supervisor['clientes_planificados'] += len(visitas_dia)

            # Pasar a siguiente ruta para próximo día
            ruta_actual_idx += 1

        # Agregar resumen
        plan_supervisor['resumen_segmentos'] = dict(resumen_segmentos)

        return plan_supervisor

    def _priorizar_rutas(self, clientes_por_ruta):
        """
        Ordena rutas según cantidad de clientes EF + Blindar
        Las rutas con más clientes prioritarios van primero
        """
        rutas_con_prioridad = []

        for ruta, clientes in clientes_por_ruta.items():
            # Contar EF + Blindar en esta ruta
            prioritarios = sum(1 for c in clientes if c['prioridad'] <= 2)
            rutas_con_prioridad.append((ruta, prioritarios))

        # Ordenar por cantidad de prioritarios (descendente)
        rutas_con_prioridad.sort(key=lambda x: x[1], reverse=True)

        return [ruta for ruta, _ in rutas_con_prioridad]

    def exportar_json(self, plan, nombre_archivo='plan_visitas.json'):
        """Exportar plan a JSON"""
        with open(nombre_archivo, 'w', encoding='utf-8') as f:
            json.dump(plan, f, indent=2, ensure_ascii=False)
        return nombre_archivo

    def exportar_excel(self, plan, nombre_archivo='plan_visitas.xlsx'):
        """Exportar plan a Excel - 6 COLUMNAS CRITICAS + adicionales"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill

            # Crear diccionario de días de visita desde CEDIS
            dias_visita_dict = {}
            for idx, row in self.df_cedis.iterrows():
                cod = int(row['Cliente: Código de Cliente'])
                dias = row['Días de visita'] if pd.notna(row['Días de visita']) else ''
                dias_visita_dict[cod] = str(dias)

            wb = Workbook()
            ws = wb.active
            ws.title = "Hoja1"

            # COLUMNAS CRITICAS (DEBEN IR EN ESTE ORDEN)
            headers = [
                "Compania sucursal",    # A - CRITICA
                "Sucursal",             # B - CRITICA
                "Codigo Zona",          # C - CRITICA
                "Ruta",                 # D - CRITICA
                "Cod Clte",             # E - CRITICA
                "Dias de visita",       # F - CRITICA
                "Fecha",                # G - Adicional
                "Dia",                  # H - Adicional
                "Clasificacion"         # I - Adicional
            ]

            # Escribir headers
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.value = header
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

            row = 2

            # Llenar datos
            for zona in sorted(plan['supervisores'].keys()):
                datos = plan['supervisores'][zona]
                for dia in datos.get('cronograma', []):
                    for visita in dia.get('visitas', []):
                        cod_cliente = int(visita.get('codigo_cliente', 0))

                        # COLUMNAS CRITICAS (6 PRIMERAS)
                        ws.cell(row=row, column=1).value = "0076|02"  # A: Compania sucursal
                        ws.cell(row=row, column=2).value = f"{str(zona).zfill(2)} - CEDIS"  # B: Sucursal
                        ws.cell(row=row, column=3).value = int(zona)  # C: Codigo Zona
                        ws.cell(row=row, column=4).value = str(visita.get('ruta', ''))  # D: Ruta
                        ws.cell(row=row, column=5).value = cod_cliente  # E: Cod Clte
                        # F: Dias de visita - BUSCAR DEL ORIGINAL CEDIS
                        ws.cell(row=row, column=6).value = dias_visita_dict.get(cod_cliente, '')

                        # COLUMNAS ADICIONALES
                        fecha_str = str(dia.get('fecha', '')).split('T')[0]
                        ws.cell(row=row, column=7).value = fecha_str  # G: Fecha
                        ws.cell(row=row, column=8).value = str(dia.get('dia_semana', ''))  # H: Dia
                        ws.cell(row=row, column=9).value = str(visita.get('clasificacion', ''))  # I: Clasificacion

                        row += 1

            # Ancho de columnas
            ws.column_dimensions['A'].width = 18
            ws.column_dimensions['B'].width = 18
            ws.column_dimensions['C'].width = 14
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 12
            ws.column_dimensions['F'].width = 25
            ws.column_dimensions['G'].width = 15
            ws.column_dimensions['H'].width = 12
            ws.column_dimensions['I'].width = 16

            # Guardar
            wb.save(nombre_archivo)
            return nombre_archivo

        except Exception as e:
            print(f"Error exportando Excel: {e}")
            return None


if __name__ == "__main__":
    generador = GeneradorPlanVisitas()

    plan = generador.generar_plan(
        mes=7,
        ano=2026,
        inicio_visitas=5,
        visitas_por_dia=20,
        dias_exclusion=[28, 29, 30, 31],
        dias_sin_visita=[5, 6]
    )

    generador.exportar_json(plan, 'plan_visitas_julio.json')
    generador.exportar_excel(plan, 'plan_visitas_julio.xlsx')

    print("Plan generado exitosamente!")
    print(f"Total de supervisores: {len(plan['supervisores'])}")
    for zona, datos in plan['supervisores'].items():
        print(f"  Zona {zona}: {datos['clientes_planificados']}/{datos['total_clientes']} clientes")
